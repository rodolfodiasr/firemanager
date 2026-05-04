import io
from datetime import datetime
from typing import Annotated
from uuid import UUID

from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import StreamingResponse
from sqlalchemy.ext.asyncio import AsyncSession

from app.api.auth import TenantContext, get_tenant_context, require_module_n2, require_module_reviewer, require_reviewer

_require_compliance    = require_module_reviewer("compliance")
_require_compliance_n2 = require_module_n2("compliance")
from app.database import get_db
from app.models.trust_score import FrameworkEnum
from app.schemas.compliance import (
    ComplianceGenerateRequest,
    ComplianceRemediateRequest,
    ComplianceReportRead,
    ComplianceReportSummary,
    FrameworkScoreItem,
    GovernanceSummary,
    TrustScoreRead,
)
from app.schemas.remediation import RemediationPlanRead
from app.services import compliance_service
from app.services import trust_score_service

router = APIRouter()


# ── CIS Benchmark reports ─────────────────────────────────────────────────────

@router.post("", response_model=ComplianceReportRead, status_code=201)
async def generate_report(
    data: ComplianceGenerateRequest,
    ctx:  Annotated[TenantContext, Depends(_require_compliance)],
    db:   Annotated[AsyncSession, Depends(get_db)],
) -> ComplianceReportRead:
    try:
        report = await compliance_service.generate_report(
            db=db,
            tenant_id=ctx.tenant.id,
            server_id=data.server_id,
            policy_id=data.policy_id,
            force_source=data.force_source,
        )
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc))
    except Exception as exc:
        raise HTTPException(status_code=500, detail=f"Erro ao gerar relatório: {exc}")
    return ComplianceReportRead.model_validate(report)


@router.get("", response_model=list[ComplianceReportSummary])
async def list_reports(
    ctx: Annotated[TenantContext, Depends(get_tenant_context)],
    db:  Annotated[AsyncSession, Depends(get_db)],
) -> list[ComplianceReportSummary]:
    reports = await compliance_service.list_reports(db, tenant_id=ctx.tenant.id)
    return [ComplianceReportSummary.model_validate(r) for r in reports]


@router.get("/{report_id}", response_model=ComplianceReportRead)
async def get_report(
    report_id: UUID,
    ctx: Annotated[TenantContext, Depends(get_tenant_context)],
    db:  Annotated[AsyncSession, Depends(get_db)],
) -> ComplianceReportRead:
    report = await compliance_service.get_report(db, ctx.tenant.id, report_id)
    if not report:
        raise HTTPException(status_code=404, detail="Relatório não encontrado")
    return ComplianceReportRead.model_validate(report)


@router.post("/{report_id}/remediate", response_model=list[RemediationPlanRead], status_code=201)
async def remediate_from_report(
    report_id: UUID,
    data: ComplianceRemediateRequest,
    ctx:  Annotated[TenantContext, Depends(_require_compliance)],
    db:   Annotated[AsyncSession, Depends(get_db)],
) -> list[RemediationPlanRead]:
    report = await compliance_service.get_report(db, ctx.tenant.id, report_id)
    if not report:
        raise HTTPException(status_code=404, detail="Relatório não encontrado")
    try:
        if data.mode == "controls":
            plans = await compliance_service.create_remediation_from_controls(
                db=db, tenant_id=ctx.tenant.id, report=report,
            )
        else:
            plans = await compliance_service.create_remediation_from_report(
                db=db, tenant_id=ctx.tenant.id, report=report,
                recommendation_index=data.recommendation_index,
            )
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc))
    except Exception as exc:
        raise HTTPException(status_code=500, detail=f"Erro ao gerar remediação: {exc}")
    return [RemediationPlanRead.model_validate(p) for p in plans]


@router.delete("/{report_id}", status_code=204)
async def delete_report(
    report_id: UUID,
    ctx: Annotated[TenantContext, Depends(_require_compliance_n2)],
    db:  Annotated[AsyncSession, Depends(get_db)],
) -> None:
    report = await compliance_service.get_report(db, ctx.tenant.id, report_id)
    if not report:
        raise HTTPException(status_code=404, detail="Relatório não encontrado")
    await db.delete(report)
    await db.flush()


@router.get("/{report_id}/export-pdf")
async def export_pdf(
    report_id: UUID,
    ctx: Annotated[TenantContext, Depends(get_tenant_context)],
    db:  Annotated[AsyncSession, Depends(get_db)],
) -> StreamingResponse:
    report = await compliance_service.get_report(db, ctx.tenant.id, report_id)
    if not report:
        raise HTTPException(status_code=404, detail="Relatório não encontrado")

    from app.models.server import Server
    from sqlalchemy import select as sa_select

    srv_result = await db.execute(sa_select(Server).where(Server.id == report.server_id))
    server = srv_result.scalar_one_or_none()
    server_name = server.name if server else str(report.server_id)

    try:
        from weasyprint import HTML

        html = _build_pdf_html(report, server_name)
        pdf_bytes = HTML(string=html).write_pdf()
        filename = f"compliance_{report.created_at.strftime('%Y%m%d_%H%M')}.pdf"
        return StreamingResponse(
            io.BytesIO(pdf_bytes),
            media_type="application/pdf",
            headers={"Content-Disposition": f'attachment; filename="{filename}"'},
        )
    except Exception as exc:
        raise HTTPException(status_code=500, detail=f"Erro ao gerar PDF: {exc}")


# ── Governance / Trust Score endpoints ────────────────────────────────────────

@router.get("/governance/scores", response_model=list[TrustScoreRead])
async def get_trust_scores(
    ctx: Annotated[TenantContext, Depends(get_tenant_context)],
    db:  Annotated[AsyncSession, Depends(get_db)],
) -> list[TrustScoreRead]:
    """Return the latest Trust Score per framework for the tenant."""
    scores = await trust_score_service.get_latest(db, ctx.tenant.id)
    return [TrustScoreRead.model_validate(s) for s in scores]


@router.post("/governance/compute", response_model=list[TrustScoreRead], status_code=201)
async def compute_trust_scores(
    ctx: Annotated[TenantContext, Depends(_require_compliance)],
    db:  Annotated[AsyncSession, Depends(get_db)],
) -> list[TrustScoreRead]:
    """Trigger Trust Score computation for all frameworks and persist the results."""
    try:
        scores = await trust_score_service.compute_all(db, ctx.tenant.id, save=True)
    except Exception as exc:
        raise HTTPException(status_code=500, detail=f"Erro ao computar Trust Score: {exc}")
    return [TrustScoreRead.model_validate(s) for s in scores]


@router.get("/governance/summary", response_model=GovernanceSummary)
async def get_governance_summary(
    ctx: Annotated[TenantContext, Depends(get_tenant_context)],
    db:  Annotated[AsyncSession, Depends(get_db)],
) -> GovernanceSummary:
    """Compact governance panel: one score per framework + Eternity narrative."""
    scores = await trust_score_service.get_latest(db, ctx.tenant.id)
    score_map = {s.framework: s for s in scores}

    def _score(key: str) -> float | None:
        s = score_map.get(key)
        return round(s.score_pct, 1) if s else None

    eternity = score_map.get(FrameworkEnum.eternity)
    return GovernanceSummary(
        eternity_score=_score(FrameworkEnum.eternity),
        cis_score=_score(FrameworkEnum.cis_benchmark),
        nist_score=_score(FrameworkEnum.nist_csf),
        iso_score=_score(FrameworkEnum.iso_27001),
        narrative=eternity.narrative if eternity else "",
        computed_at=eternity.computed_at if eternity else None,
        scores=[TrustScoreRead.model_validate(s) for s in scores],
    )


@router.get("/governance/history/{framework}", response_model=list[FrameworkScoreItem])
async def get_score_history(
    framework: str,
    ctx:   Annotated[TenantContext, Depends(get_tenant_context)],
    db:    Annotated[AsyncSession, Depends(get_db)],
    limit: int = Query(default=30, ge=1, le=100),
) -> list[FrameworkScoreItem]:
    """Historical trend for a single framework (most recent first)."""
    try:
        fw = FrameworkEnum(framework)
    except ValueError:
        raise HTTPException(
            status_code=400,
            detail=f"Framework inválido. Valores aceitos: {[e.value for e in FrameworkEnum]}",
        )
    scores = await trust_score_service.get_history(db, ctx.tenant.id, fw, limit=limit)
    return [FrameworkScoreItem.model_validate(s) for s in scores]


@router.get("/governance/export-excel")
async def export_governance_excel(
    ctx: Annotated[TenantContext, Depends(get_tenant_context)],
    db:  Annotated[AsyncSession, Depends(get_db)],
) -> StreamingResponse:
    """Export full governance evidence workbook (xlsx) for auditors."""
    try:
        import openpyxl
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter
    except ImportError:
        raise HTTPException(status_code=500, detail="openpyxl não instalado")

    scores = await trust_score_service.get_latest(db, ctx.tenant.id)
    reports = await compliance_service.list_reports(db, ctx.tenant.id)
    score_map = {s.framework: s for s in scores}

    wb = openpyxl.Workbook()

    # ── Sheet 1: Eternity Trust Score ─────────────────────────────────────────
    ws = wb.active
    ws.title = "Eternity Trust Score"

    _hdr(ws, "A1", "FireManager — Eternity Trust Score", bold=True, size=16)
    _hdr(ws, "A2", f"Gerado em: {datetime.utcnow().strftime('%d/%m/%Y %H:%M')} UTC")

    eternity = score_map.get(FrameworkEnum.eternity)
    ws["A4"] = "Eternity Trust Score"
    ws["B4"] = eternity.score_pct if eternity else "N/A"
    ws["A5"] = "CIS Benchmark"
    ws["B5"] = score_map[FrameworkEnum.cis_benchmark].score_pct if FrameworkEnum.cis_benchmark in score_map else "N/A"
    ws["A6"] = "NIST CSF"
    ws["B6"] = score_map[FrameworkEnum.nist_csf].score_pct if FrameworkEnum.nist_csf in score_map else "N/A"
    ws["A7"] = "ISO 27001"
    ws["B7"] = score_map[FrameworkEnum.iso_27001].score_pct if FrameworkEnum.iso_27001 in score_map else "N/A"

    ws["A9"] = "Narrativa Executiva"
    ws["A9"].font = Font(bold=True)
    ws["A10"] = eternity.narrative if eternity else "Nenhum score computado. Execute POST /compliance/governance/compute."
    ws["A10"].alignment = Alignment(wrap_text=True)
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 15
    ws.row_dimensions[10].height = 120

    # ── Sheet 2: Framework Breakdown ─────────────────────────────────────────
    ws2 = wb.create_sheet("Framework Breakdown")
    headers = ["Framework", "Score (%)", "Computed At", "Breakdown (JSON)"]
    for col, h in enumerate(headers, 1):
        cell = ws2.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="E85D04")
        cell.font = Font(bold=True, color="FFFFFF")

    for row_n, s in enumerate(scores, 2):
        import json
        ws2.cell(row=row_n, column=1, value=s.framework)
        ws2.cell(row=row_n, column=2, value=round(s.score_pct, 1))
        ws2.cell(row=row_n, column=3, value=s.computed_at.strftime("%d/%m/%Y %H:%M"))
        ws2.cell(row=row_n, column=4, value=json.dumps(s.breakdown, ensure_ascii=False)[:500])

    for col in range(1, 5):
        ws2.column_dimensions[get_column_letter(col)].width = 25

    # ── Sheet 3: CIS Compliance Reports ──────────────────────────────────────
    ws3 = wb.create_sheet("CIS Reports")
    cis_headers = [
        "ID", "Server ID", "Source", "Policy", "Framework",
        "Score (%)", "Passed", "Failed", "N/A", "Total", "Date",
    ]
    for col, h in enumerate(cis_headers, 1):
        cell = ws3.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="2D3748")
        cell.font = Font(bold=True, color="FFFFFF")

    for row_n, r in enumerate(reports, 2):
        ws3.cell(row=row_n, column=1,  value=str(r.id))
        ws3.cell(row=row_n, column=2,  value=str(r.server_id))
        ws3.cell(row=row_n, column=3,  value=r.source)
        ws3.cell(row=row_n, column=4,  value=r.policy_name)
        ws3.cell(row=row_n, column=5,  value=r.framework)
        ws3.cell(row=row_n, column=6,  value=round(r.score_pct, 1))
        ws3.cell(row=row_n, column=7,  value=r.passed)
        ws3.cell(row=row_n, column=8,  value=r.failed)
        ws3.cell(row=row_n, column=9,  value=r.not_applicable)
        ws3.cell(row=row_n, column=10, value=r.total_checks)
        ws3.cell(row=row_n, column=11, value=r.created_at.strftime("%d/%m/%Y %H:%M"))

    for col in range(1, len(cis_headers) + 1):
        ws3.column_dimensions[get_column_letter(col)].width = 18

    # ── Sheet 4: ISO 27001 Controls ───────────────────────────────────────────
    iso = score_map.get(FrameworkEnum.iso_27001)
    ws4 = wb.create_sheet("ISO 27001")
    _hdr(ws4, "A1", "ISO 27001 — Evidência de Conformidade", bold=True)
    _hdr(ws4, "A2", "Baseado nos controles de Security Hardening P1–P6")

    iso_headers = ["Controle ISO 27001", "Score (%)", "Fase FireManager"]
    for col, h in enumerate(iso_headers, 1):
        cell = ws4.cell(row=4, column=col, value=h)
        cell.font = Font(bold=True)

    iso_controls_map = {
        "A.10_cryptography":   ("A.10 Criptografia", "P1 — Credential Encryption"),
        "A.9_access_control":  ("A.9 Controle de Acesso", "P2 — Row-Level Security"),
        "A.8.15_logging":      ("A.8.15 Logging", "P3 — Audit Log Immutability"),
        "A.8.6_operations":    ("A.8.6 Operações", "P4/P5 — Operation Workflow"),
        "A.9.4_system_access": ("A.9.4 Acesso ao Sistema", "P6 — Multi-Sig Approval"),
        "A.12_compliance":     ("A.12 Conformidade Técnica", "CIS Benchmark"),
    }
    iso_bd = iso.breakdown.get("iso_controls", {}) if iso else {}
    for row_n, (key, (label, phase)) in enumerate(iso_controls_map.items(), 5):
        ws4.cell(row=row_n, column=1, value=label)
        ws4.cell(row=row_n, column=2, value=iso_bd.get(key, "N/A"))
        ws4.cell(row=row_n, column=3, value=phase)

    for col in range(1, 4):
        ws4.column_dimensions[get_column_letter(col)].width = 35

    # ── Sheet 5: NIST CSF Functions ───────────────────────────────────────────
    nist = score_map.get(FrameworkEnum.nist_csf)
    ws5 = wb.create_sheet("NIST CSF")
    _hdr(ws5, "A1", "NIST Cybersecurity Framework — Funções", bold=True)

    nist_headers = ["Função NIST CSF", "Score (%)"]
    for col, h in enumerate(nist_headers, 1):
        cell = ws5.cell(row=3, column=col, value=h)
        cell.font = Font(bold=True)

    nist_funcs = {
        "identify": "Identificar (ID)",
        "protect":  "Proteger (PR)",
        "detect":   "Detectar (DE)",
        "respond":  "Responder (RS)",
        "recover":  "Recuperar (RC)",
    }
    nist_bd = nist.breakdown.get("nist_functions", {}) if nist else {}
    for row_n, (key, label) in enumerate(nist_funcs.items(), 4):
        ws5.cell(row=row_n, column=1, value=label)
        ws5.cell(row=row_n, column=2, value=nist_bd.get(key, "N/A"))

    ws5.column_dimensions["A"].width = 30
    ws5.column_dimensions["B"].width = 15

    # ── Serialize ─────────────────────────────────────────────────────────────
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f"governance_{datetime.utcnow().strftime('%Y%m%d_%H%M')}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


# ── PDF helper ────────────────────────────────────────────────────────────────

def _hdr(ws, cell: str, value: str, bold: bool = False, size: int = 12) -> None:
    ws[cell] = value
    ws[cell].font = Font(bold=bold, size=size)


def _build_pdf_html(report, server_name: str) -> str:
    score_color = (
        "#e53e3e" if report.score_pct < 50
        else "#d69e2e" if report.score_pct < 75
        else "#38a169"
    )

    source_label = "Wazuh SCA (agente)" if report.source == "wazuh" else "SSH (snapshot)"
    created = report.created_at.strftime("%d/%m/%Y %H:%M")

    failed_controls = [c for c in (report.controls or []) if c.get("result") == "failed"][:20]
    controls_rows = ""
    for ctrl in failed_controls:
        risk = ctrl.get("risk_level", "medium")
        risk_color = {
            "critical": "#e53e3e", "high": "#dd6b20",
            "medium": "#d69e2e", "low": "#718096",
        }.get(risk, "#718096")
        controls_rows += f"""
        <tr>
          <td style="color:#555;font-size:11px">{ctrl.get('control_id','')}</td>
          <td style="font-size:12px">{ctrl.get('title','')}</td>
          <td style="color:{risk_color};font-weight:bold;font-size:11px;text-transform:uppercase">{risk}</td>
          <td style="font-size:11px;color:#333">{ctrl.get('remediation','')[:120]}</td>
        </tr>"""

    recs_html = ""
    for rec in (report.ai_recommendations or [])[:10]:
        recs_html += f"""
        <div class="rec-item">
          <span class="rec-priority">#{rec.get('priority','')}</span>
          <strong>{rec.get('title','')}</strong>
          <p style="margin:4px 0 4px 0;color:#555;font-size:13px">{rec.get('description','')}</p>
          <pre class="code">{rec.get('remediation_steps','')}</pre>
        </div>"""

    ai_summary_html = (report.ai_summary or "").replace("\n", "<br>")

    return f"""<!DOCTYPE html>
<html lang="pt-BR">
<head>
  <meta charset="UTF-8">
  <style>
    body {{ font-family: Arial, sans-serif; margin: 40px; color: #1a1a1a; line-height: 1.5; }}
    .header {{ border-bottom: 3px solid #e85d04; padding-bottom: 16px; margin-bottom: 24px; }}
    .header h1 {{ color: #e85d04; margin: 0 0 4px 0; font-size: 22px; }}
    .meta {{ color: #666; font-size: 13px; }}
    .score-box {{ display: inline-block; background: {score_color}; color: white; font-size: 36px;
      font-weight: bold; padding: 12px 24px; border-radius: 8px; margin: 8px 0; }}
    .counters {{ display: flex; gap: 24px; margin: 12px 0 24px; }}
    .counter {{ text-align: center; }}
    .counter .num {{ font-size: 24px; font-weight: bold; }}
    .counter .lbl {{ font-size: 11px; color: #888; text-transform: uppercase; }}
    .section-label {{ font-size: 11px; font-weight: bold; color: #888; text-transform: uppercase;
      letter-spacing: 0.05em; margin: 20px 0 8px; }}
    .summary-box {{ background: #f9f9f9; border: 1px solid #e0e0e0; border-radius: 6px;
      padding: 16px; font-size: 14px; margin-bottom: 24px; }}
    table {{ width: 100%; border-collapse: collapse; font-size: 12px; margin-bottom: 24px; }}
    th {{ background: #f0f0f0; text-align: left; padding: 8px; font-size: 11px; color: #555;
      text-transform: uppercase; }}
    td {{ padding: 7px 8px; border-bottom: 1px solid #eee; vertical-align: top; }}
    .rec-item {{ background: #fff; border: 1px solid #e0e0e0; border-radius: 6px;
      padding: 12px 16px; margin-bottom: 10px; }}
    .rec-priority {{ display: inline-block; background: #e85d04; color: white; font-size: 11px;
      font-weight: bold; padding: 1px 6px; border-radius: 10px; margin-right: 6px; }}
    .code {{ background: #1a1a2e; color: #a8ff78; font-size: 11px; padding: 8px 12px;
      border-radius: 4px; overflow: hidden; margin: 6px 0 0; font-family: monospace; }}
    .footer {{ border-top: 1px solid #eee; padding-top: 12px; font-size: 11px; color: #aaa; margin-top: 32px; }}
    .badge {{ display: inline-block; font-size: 11px; padding: 2px 8px; border-radius: 10px;
      background: #e8f4fd; color: #1a6fa0; }}
  </style>
</head>
<body>
  <div class="header">
    <h1>FireManager — Relatório de Conformidade {report.framework.upper()}</h1>
    <p class="meta">
      Servidor: <strong>{server_name}</strong> &nbsp;|&nbsp;
      Política: <strong>{report.policy_name}</strong> &nbsp;|&nbsp;
      Fonte: <span class="badge">{source_label}</span> &nbsp;|&nbsp;
      Gerado em {created}
    </p>
  </div>

  <div class="section-label">Score de Conformidade</div>
  <div class="score-box">{report.score_pct:.1f}%</div>
  <div class="counters">
    <div class="counter"><div class="num" style="color:#38a169">{report.passed}</div><div class="lbl">Passou</div></div>
    <div class="counter"><div class="num" style="color:#e53e3e">{report.failed}</div><div class="lbl">Falhou</div></div>
    <div class="counter"><div class="num" style="color:#718096">{report.not_applicable}</div><div class="lbl">N/A</div></div>
    <div class="counter"><div class="num">{report.total_checks}</div><div class="lbl">Total</div></div>
  </div>

  <div class="section-label">Resumo Executivo</div>
  <div class="summary-box">{ai_summary_html}</div>

  <div class="section-label">Controles Reprovados (top 20)</div>
  <table>
    <tr><th>ID</th><th>Controle</th><th>Risco</th><th>Remediação</th></tr>
    {controls_rows if controls_rows else '<tr><td colspan="4" style="color:#aaa;text-align:center">Nenhuma falha encontrada</td></tr>'}
  </table>

  <div class="section-label">Recomendações Prioritárias</div>
  {recs_html if recs_html else '<p style="color:#aaa;font-size:13px">Nenhuma recomendação gerada.</p>'}

  <div class="footer">
    FireManager v0.1.0 &nbsp;|&nbsp; Relatório ID: {report.id}
  </div>
</body>
</html>"""
